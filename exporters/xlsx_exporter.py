#!/usr/bin/env python3
"""
XLSX Exporter for YMYL Audit Tool

Converts markdown reports to Excel spreadsheets with multiple worksheets and formatting.
"""

import io
import re
import json
from datetime import datetime
from typing import Optional, Dict, List, Any
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - XLSX export will be disabled")
from utils.logging_utils import setup_logger

logger = setup_logger(__name__)


class XLSXExporter:
    """
    Converts markdown reports to Excel spreadsheets with multiple worksheets.
    
    Features:
    - Summary worksheet with key metrics
    - Violations worksheet with detailed findings
    - Content sections worksheet
    - Professional formatting with colors
    - Auto-sized columns
    """
    
    def __init__(self):
        """Initialize the XLSX exporter."""
        self.workbook = None
        self.colors = {
            'critical': 'FFE74C3C',  # Red
            'high': 'FFE67E22',      # Orange  
            'medium': 'FFF39C12',    # Yellow
            'low': 'FF3498DB',       # Blue
            'header': 'FF2C3E50',    # Dark blue
            'success': 'FF27AE60',   # Green
            'background': 'FFF8F9FA'  # Light gray
        }
        logger.info("XLSXExporter initialized")

    def convert(self, markdown_content: str, title: str = "YMYL Compliance Audit Report") -> bytes:
        """
        Convert markdown content to Excel spreadsheet.
        
        Args:
            markdown_content (str): Markdown content to convert
            title (str): Document title
            
        Returns:
            bytes: Excel document as bytes
        """
        try:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl library is required for XLSX export")
                
            logger.info(f"Converting markdown to XLSX ({len(markdown_content):,} characters)")
            
            # Parse markdown content
            parsed_data = self._parse_markdown_content(markdown_content)
            
            # Create workbook
            self.workbook = Workbook()
            
            # Remove default sheet
            self.workbook.remove(self.workbook.active)
            
            # Create worksheets with only what exists in the report
            self._create_summary_worksheet(parsed_data, title)
            self._create_violations_worksheet(parsed_data)
            self._create_content_sections_worksheet(parsed_data)
            self._create_raw_data_worksheet(markdown_content)
            
            # Save to memory
            excel_buffer = io.BytesIO()
            self.workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info("XLSX conversion successful")
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"XLSX conversion error: {e}")
            return self._create_error_workbook(str(e))

    def _parse_markdown_content(self, markdown_content: str) -> Dict[str, Any]:
        """
        Parse markdown content to extract exactly what's in the existing report.
        
        Args:
            markdown_content (str): Markdown content
            
        Returns:
            dict: Parsed data structure
        """
        try:
            data = {
                'title': '',
                'date': '',
                'sections': [],
                'violations': [],
                'processing_info': []
            }
            
            lines = markdown_content.split('\n')
            current_section = None
            in_processing_summary = False
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Extract title
                if line.startswith('# ') and not data['title']:
                    data['title'] = line[2:]
                
                # Extract date
                elif line.startswith('**Date:**'):
                    data['date'] = line.replace('**Date:**', '').strip()
                
                # Track sections
                elif line.startswith('## '):
                    section_name = line[3:]
                    current_section = {
                        'name': section_name,
                        'violations': []
                    }
                    data['sections'].append(current_section)
                    
                    # Check if this is processing summary
                    in_processing_summary = 'processing summary' in section_name.lower()
                
                # Extract violations with severity indicators
                elif any(indicator in line for indicator in ['🔴', '🟠', '🟡', '🔵']):
                    violation = self._parse_violation_line(line)
                    if violation:
                        violation['section'] = current_section['name'] if current_section else 'Unknown Section'
                        data['violations'].append(violation)
                        if current_section:
                            current_section['violations'].append(violation)
                
                # Extract processing info from processing summary
                elif in_processing_summary and (line.startswith('- ') or line.startswith('**')):
                    data['processing_info'].append(line)
            
            return data
            
        except Exception as e:
            logger.error(f"Error parsing markdown content: {e}")
            return {
                'title': 'Parse Error',
                'date': datetime.now().strftime("%Y-%m-%d"),
                'sections': [],
                'violations': [],
                'processing_info': []
            }

    def _parse_violation_line(self, line: str) -> Optional[Dict[str, str]]:
        """Parse a violation line and extract components."""
        try:
            # Map severity indicators
            severity_map = {
                '🔴': 'Critical',
                '🟠': 'High', 
                '🟡': 'Medium',
                '🔵': 'Low'
            }
            
            severity = 'Unknown'
            for indicator, sev in severity_map.items():
                if indicator in line:
                    severity = sev
                    line = line.replace(indicator, '').strip()
                    break
            
            # Extract violation text (everything after severity indicator)
            violation_text = line
            
            # Try to extract more structured data if available
            issue_match = re.search(r'\*\*Issue:\*\*\s*([^*]+)', line)
            text_match = re.search(r'\*\*Problematic Text:\*\*\s*"([^"]*)"', line)
            fix_match = re.search(r'\*\*Suggested Fix:\*\*\s*"([^"]*)"', line)
            
            return {
                'severity': severity,
                'full_text': violation_text,
                'issue': issue_match.group(1).strip() if issue_match else violation_text,
                'problematic_text': text_match.group(1) if text_match else '',
                'suggested_fix': fix_match.group(1) if fix_match else ''
            }
            
        except Exception as e:
            logger.warning(f"Error parsing violation line: {e}")
            return None

    def _extract_summary_stats(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract only the actual summary statistics from the existing report."""
        try:
            total_violations = len(data['violations'])
            critical_count = len([v for v in data['violations'] if v['severity'] == 'Critical'])
            high_count = len([v for v in data['violations'] if v['severity'] == 'High'])
            medium_count = len([v for v in data['violations'] if v['severity'] == 'Medium'])
            low_count = len([v for v in data['violations'] if v['severity'] == 'Low'])
            
            sections_with_violations = len([s for s in data['sections'] if s['violations']])
            total_sections = len([s for s in data['sections'] if not 'processing summary' in s['name'].lower()])
            
            return {
                'total_violations': total_violations,
                'critical_violations': critical_count,
                'high_violations': high_count,
                'medium_violations': medium_count,
                'low_violations': low_count,
                'sections_analyzed': total_sections,
                'sections_with_violations': sections_with_violations
            }
            
        except Exception as e:
            logger.error(f"Error extracting summary stats: {e}")
            return {}

    def _create_summary_worksheet(self, data: Dict[str, Any], title: str):
        """Create summary worksheet with only actual metrics from the report."""
        try:
            ws = self.workbook.create_sheet("Summary")
            
            # Title
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True, color='FFFFFFFF')
            ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            ws.merge_cells('A1:B1')
            
            # Date
            ws['A3'] = "Report Date:"
            ws['B3'] = data.get('date', datetime.now().strftime("%Y-%m-%d"))
            ws['A3'].font = Font(bold=True)
            
            # Only add statistics that actually exist
            stats = self._extract_summary_stats(data)
            row = 5
            
            metrics = [
                ("Total Violations:", stats.get('total_violations', 0)),
                ("Critical Issues:", stats.get('critical_violations', 0)),
                ("High Priority Issues:", stats.get('high_violations', 0)),
                ("Medium Priority Issues:", stats.get('medium_violations', 0)),
                ("Low Priority Issues:", stats.get('low_violations', 0)),
                ("Sections Analyzed:", stats.get('sections_analyzed', 0)),
                ("Sections with Issues:", stats.get('sections_with_violations', 0))
            ]
            
            for metric_name, metric_value in metrics:
                ws[f'A{row}'] = metric_name
                ws[f'B{row}'] = metric_value
                ws[f'A{row}'].font = Font(bold=True)
                
                # Color code only critical and high severity
                if 'Critical' in metric_name and metric_value > 0:
                    ws[f'B{row}'].fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
                elif 'High' in metric_name and metric_value > 0:
                    ws[f'B{row}'].fill = PatternFill(start_color=self.colors['high'], end_color=self.colors['high'], fill_type='solid')
                
                row += 1
            
            # Add processing summary if it exists
            if data.get('processing_info'):
                row += 2
                ws[f'A{row}'] = "Processing Summary:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for info in data['processing_info']:
                    ws[f'A{row}'] = info
                    row += 1
            
            # Auto-size columns
            self._auto_size_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating summary worksheet: {e}")

    def _create_violations_worksheet(self, data: Dict[str, Any]):
        """Create violations worksheet with exactly what's in the report."""
        try:
            ws = self.workbook.create_sheet("Violations")
            
            # Headers based on actual violation structure
            headers = ["Section", "Severity", "Violation Text"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFFFF')
                cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Add violations exactly as they appear in report
            row = 2
            for violation in data.get('violations', []):
                ws.cell(row=row, column=1, value=violation.get('section', 'Unknown'))
                ws.cell(row=row, column=2, value=violation['severity'])
                ws.cell(row=row, column=3, value=violation['full_text'])
                
                # Color code severity
                severity_color = self.colors.get(violation['severity'].lower(), self.colors['background'])
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color=severity_color, end_color=severity_color, fill_type='solid')
                
                row += 1
            
            # Add formatting
            self._add_table_formatting(ws, row - 1, 3)
            self._auto_size_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating violations worksheet: {e}")

    def _create_content_sections_worksheet(self, data: Dict[str, Any]):
        """Create content sections overview based on actual report sections."""
        try:
            ws = self.workbook.create_sheet("Content Sections")
            
            # Headers
            headers = ["Section Name", "Violations Count", "Status"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFFFF')
                cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Add sections data (exclude processing summary)
            row = 2
            for section in data.get('sections', []):
                # Skip processing summary section
                if 'processing summary' in section['name'].lower():
                    continue
                    
                violations = section.get('violations', [])
                violation_count = len(violations)
                status = 'Issues Found' if violation_count > 0 else 'Clean'
                
                ws.cell(row=row, column=1, value=section['name'])
                ws.cell(row=row, column=2, value=violation_count)
                ws.cell(row=row, column=3, value=status)
                
                # Color code status
                status_color = self.colors['critical'] if violation_count > 0 else self.colors['success']
                ws.cell(row=row, column=3).fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                
                row += 1
            
            # Add formatting
            self._add_table_formatting(ws, row - 1, 3)
            self._auto_size_columns(ws)
            
        except Exception as e:
            logger.error(f"Error creating content sections worksheet: {e}")

    def _create_raw_data_worksheet(self, markdown_content: str):
        """Create raw data worksheet with full markdown content."""
        try:
            ws = self.workbook.create_sheet("Raw Report")
            
            # Add header
            ws['A1'] = "Full Markdown Report"
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFFFF')
            ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
            
            # Split content into manageable chunks for Excel cells
            lines = markdown_content.split('\n')
            row = 3
            
            for line in lines:
                if len(line) > 32767:  # Excel cell limit
                    # Split very long lines
                    chunks = [line[i:i+32767] for i in range(0, len(line), 32767)]
                    for chunk in chunks:
                        ws.cell(row=row, column=1, value=chunk)
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=line)
                    row += 1
            
            # Set column width
            ws.column_dimensions['A'].width = 100
            
        except Exception as e:
            logger.error(f"Error creating raw data worksheet: {e}")

    def _add_table_formatting(self, ws: Worksheet, max_row: int, max_col: int):
        """Add professional table formatting."""
        try:
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
        except Exception as e:
            logger.error(f"Error adding table formatting: {e}")

    def _auto_size_columns(self, ws: Worksheet):
        """Auto-size columns based on content."""
        try:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.error(f"Error auto-sizing columns: {e}")

    def _create_error_workbook(self, error_message: str) -> bytes:
        """Create error workbook when conversion fails."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Error"
            
            ws['A1'] = "Export Error"
            ws['A1'].font = Font(size=16, bold=True, color='FFFFFFFF')
            ws['A1'].fill = PatternFill(start_color=self.colors['critical'], end_color=self.colors['critical'], fill_type='solid')
            
            ws['A3'] = f"Failed to convert report: {error_message}"
            ws['A5'] = "Please try again or contact support if the problem persists."
            
            error_buffer = io.BytesIO()
            wb.save(error_buffer)
            error_buffer.seek(0)
            return error_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating error workbook: {e}")
            return b"Error creating Excel document"

    def validate_markdown(self, markdown_content: str) -> bool:
        """Validate markdown content before conversion."""
        try:
            if not markdown_content or not markdown_content.strip():
                return False
            
            if len(markdown_content) > 10000000:  # 10MB limit
                logger.warning(f"Markdown content very large: {len(markdown_content):,} characters")
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"Markdown validation error: {e}")
            return False

    def get_document_info(self, markdown_content: str) -> Dict[str, Any]:
        """Get information about the Excel file based on actual report content."""
        try:
            parsed_data = self._parse_markdown_content(markdown_content)
            stats = self._extract_summary_stats(parsed_data)
            
            return {
                'total_violations': stats.get('total_violations', 0),
                'total_sections': stats.get('sections_analyzed', 0),
                'critical_violations': stats.get('critical_violations', 0),
                'worksheets': ['Summary', 'Violations', 'Content Sections', 'Raw Report'],
                'estimated_size': f"{len(markdown_content) // 1024}KB"
            }
            
        except Exception as e:
            logger.error(f"Error getting document info: {e}")
            return {
                'total_violations': 0,
                'total_sections': 0,
                'critical_violations': 0,
                'worksheets': ['Error'],
                'estimated_size': '0KB'
            }